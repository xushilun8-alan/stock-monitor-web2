#!/usr/bin/env python3
"""
阶段买入工具函数 (app/stage_buying/utils.py)
飞书通知 + Excel 导入导出
"""

import os
import json
import subprocess
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional
from io import BytesIO

FEISHU_TARGET = "ou_268fcd21ee877df7e4d16305a4892d7c"

# 通知状态文件
_NOTIF_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'stage_notif_status.json')


def _ensure_data_dir():
    os.makedirs(os.path.dirname(_NOTIF_FILE), exist_ok=True)


def _is_stage_notified(stock_id: int, stage_id: int) -> bool:
    key = f"{stock_id}_{stage_id}"
    _ensure_data_dir()
    if not os.path.exists(_NOTIF_FILE):
        return False
    try:
        with open(_NOTIF_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get(key, False)
    except Exception:
        return False


def _mark_stage_notified(stock_id: int, stage_id: int) -> None:
    key = f"{stock_id}_{stage_id}"
    _ensure_data_dir()
    try:
        data = {}
        if os.path.exists(_NOTIF_FILE):
            with open(_NOTIF_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
        data[key] = True
        with open(_NOTIF_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def send_stage_trigger_notification(
    stock_name: str,
    stock_code: str,
    current_price: float,
    stage_num: int,
    buy_price: float,
    shares: int,
    trigger_time: str,
) -> bool:
    """发送阶段买入触发飞书通知"""
    if _is_stage_notified(int(stock_code.split('_')[-1]) if not stock_code.isdigit() else 0, stage_num):
        return False

    message = f"""📉 股票阶段买入提醒
股票名称：{stock_name} ({stock_code})
当前价格：{current_price}
触发阶段：第{stage_num}阶段
阶段买入价：{buy_price}
建议买入股数：{shares}
触发时间：{trigger_time}"""

    cmd = [
        'openclaw', 'message', 'send',
        '--channel', 'feishu',
        '--target', FEISHU_TARGET,
        '--message', message
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
        return result.returncode == 0
    except Exception:
        return False


def send_test_notification() -> bool:
    """发送测试通知"""
    message = f"""✅【阶段买入测试】
时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
阶段买入策略跟踪系统，飞书通知功能正常！"""

    cmd = [
        'openclaw', 'message', 'send',
        '--channel', 'feishu',
        '--target', FEISHU_TARGET,
        '--message', message
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
        return result.returncode == 0
    except Exception:
        return False


# ── Excel 导入导出 ────────────────────────────────────────

def export_to_excel(stocks: List[Dict[str, Any]]) -> BytesIO:
    """
    导出股票列表及阶段详情到 Excel
    返回 BytesIO 对象
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除默认 sheet

    # Sheet 1: 股票汇总
    ws_sum = wb.create_sheet('股票汇总')
    ws_sum.append([
        '代码', '名称', '初始单价', '初始股数', '每阶股数', '总阶段数',
        '序号系数', '幅度系数', '跌幅系数', '目标价基准', '最小幅度',
        '幅度乘数', '底价', '当前价格', '当前阶段', '已执行/总阶段',
        '总买入金额', '已执行金额', '创建时间'
    ])
    for s in stocks:
        ws_sum.append([
            s.get('code', ''),
            s.get('name', ''),
            s.get('initial_price', ''),
            s.get('initial_shares', ''),
            s.get('per_stage_shares', ''),
            s.get('stage_count', ''),
            s.get('serial_coefficient', ''),
            s.get('amplitude_coefficient', ''),
            s.get('decline_coefficient', ''),
            s.get('target_price', ''),
            s.get('min_amplitude', ''),
            s.get('amplitude_multiplier', ''),
            s.get('floor_price', ''),
            s.get('current_price', ''),
            s.get('current_stage', ''),
            f"{s.get('executed_count', 0)}/{s.get('total_stages', 0)}",
            s.get('total_investment', ''),
            s.get('executed_investment', ''),
            s.get('created_at', ''),
        ])

    # Sheet 2: 阶段详情
    ws_stage = wb.create_sheet('阶段详情')
    ws_stage.append([
        '股票代码', '股票名称', '阶段序号', '幅度', '买入单价', '股数',
        '买入金额', '底价亏损', '亏损率(%)', '目标收益', '期望收益',
        '收益率(%)', '状态'
    ])
    for s in stocks:
        for st in s.get('stages', []):
            ws_stage.append([
                s.get('code', ''),
                s.get('name', ''),
                st.get('stage_number', ''),
                st.get('amplitude', ''),
                st.get('buy_price', ''),
                st.get('shares', ''),
                st.get('buy_amount', ''),
                st.get('floor_loss', ''),
                st.get('loss_rate', ''),
                st.get('target_income', ''),
                st.get('expected_return', ''),
                st.get('return_rate', ''),
                st.get('status', ''),
            ])

    # 设置列宽
    for ws in [ws_sum, ws_stage]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_from_excel(file_stream) -> List[Dict[str, Any]]:
    """
    从 Excel 文件导入股票数据
    返回股票参数列表
    """
    wb = openpyxl.load_workbook(file_stream)
    ws = wb['股票汇总'] if '股票汇总' in wb.sheetnames else wb.active

    stocks = []
    headers = [cell.value for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers) if h}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            stock = {
                'code': str(row[header_map.get('代码', 0)] or '').strip(),
                'name': str(row[header_map.get('名称', 1)] or '').strip(),
                'initial_price': float(row[header_map.get('初始单价', 2)] or 0),
                'initial_shares': int(row[header_map.get('初始股数', 3)] or 0),
                'per_stage_shares': int(row[header_map.get('每阶股数', 4)] or 0),
                'stage_count': int(row[header_map.get('总阶段数', 5)] or 9),
                'serial_coefficient': float(row[header_map.get('序号系数', 6)] or 1.0),
                'amplitude_coefficient': float(row[header_map.get('幅度系数', 7)] or 0.08),
                'decline_coefficient': float(row[header_map.get('跌幅系数', 8)] or 0.975),
                'target_price': float(row[header_map.get('目标价基准', 9)]) if row[header_map.get('目标价基准', 9)] else None,
                'min_amplitude': float(row[header_map.get('最小幅度', 10)] or 0.98),
                'amplitude_multiplier': float(row[header_map.get('幅度乘数', 11)] or 1.001),
                'floor_price': float(row[header_map.get('底价', 12)]) if row[header_map.get('底价', 12)] else None,
            }
            stocks.append(stock)
        except (ValueError, TypeError, IndexError):
            continue

    return stocks
